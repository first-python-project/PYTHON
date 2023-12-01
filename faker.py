import openpyxl
from faker import Faker
import os

upload_path = "uploads"
workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = "이름"
worksheet['B1'] = "전화번호"
worksheet['C1'] = "우편번호"
worksheet['D1'] = "주소"
worksheet['E1'] = "이메일"

fake = Faker('ko_KR')


for row in range(2, 50): #2행부터 시작해 50줄까지 생성
    worksheet.cell(row=row, column=1, value=fake.name())
    worksheet.cell(row=row, column=2, value=fake.phone_number())
    worksheet.cell(row=row, column=3, value=fake.postcode())
    worksheet.cell(row=row, column=4, value=fake.address())
    worksheet.cell(row=row, column=5, value=fake.email())

file_path = os.path.join(upload_path, "customer_list.xlsx")
workbook.save(file_path)
workbook.save("customer_list.xlsx")

